"""Idempotently ingest a flat, explicitly mapped catalogue workbook."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.dialects.postgresql import insert

from app.db.models import (
    Brand,
    Variant,
    VariantComfortSpec,
    VariantConnectedSpec,
    VariantInfotainmentSpec,
    VariantLightingSpec,
    VariantPhysicalSpec,
    VariantPowertrainSpec,
    VariantPrice,
    VariantSafetySpec,
    VehicleModel,
)
from app.db.session import SessionLocal

REQUIRED = {
    "catalogue_id",
    "model",
    "trim",
    "variant_name",
    "model_year",
    "city",
    "ex_showroom_price",
    "fuel_type",
    "transmission_type",
    "drivetrain",
}
SPEC_MODELS = (
    VariantPowertrainSpec,
    VariantPhysicalSpec,
    VariantSafetySpec,
    VariantComfortSpec,
    VariantInfotainmentSpec,
    VariantLightingSpec,
    VariantConnectedSpec,
)


def _key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def _value(value: object) -> object:
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped or stripped.lower() in {"na", "n/a", "null", "unknown", "-"}:
        return None
    lowered = stripped.lower()
    if lowered in {"yes", "true", "available"}:
        return True
    if lowered in {"no", "false", "not available"}:
        return False
    return lowered.replace(" ", "_") if lowered in {
        "petrol", "cng", "hybrid", "electric", "manual", "automatic", "amt",
        "torque converter", "e cvt", "fwd", "rwd", "awd", "4wd",
    } else stripped


def _upsert_spec(session: Any, model: Any, variant_id: int, row: dict[str, object]) -> None:
    columns = {column.name for column in model.__table__.columns} - {"variant_id"}
    values = {name: row[name] for name in columns if row.get(name) is not None}
    if not values and model is not VariantPowertrainSpec:
        return
    statement = insert(model).values(variant_id=variant_id, **values)
    session.execute(
        statement.on_conflict_do_update(
            index_elements=["variant_id"],
            set_=values,
        )
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument(
        "--mapping",
        type=Path,
        help="JSON object mapping source headers to canonical database column names",
    )
    parser.add_argument("--headers", action="store_true", help="Print normalized headers and exit")
    parser.add_argument("--deactivate-missing", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    rows = sheet.iter_rows(values_only=True)
    source_headers = [str(value or "").strip() for value in next(rows)]
    if args.headers:
        print(json.dumps({_key(value): value for value in source_headers}, indent=2))
        return
    mapping = (
        json.loads(args.mapping.read_text(encoding="utf-8")) if args.mapping else {}
    )
    headers = [_key(mapping.get(header, header)) for header in source_headers]
    missing = REQUIRED - set(headers)
    if missing:
        raise SystemExit(
            "Workbook mapping is incomplete. Missing canonical columns: "
            + ", ".join(sorted(missing))
        )
    known = REQUIRED | {"brand", "on_road_price", "is_active"}
    for model in SPEC_MODELS:
        known |= {column.name for column in model.__table__.columns}
    unknown = sorted(set(headers) - known)
    if unknown:
        print("Ignored unmapped columns: " + ", ".join(unknown))

    seen: set[str] = set()
    ingested = 0
    with SessionLocal.begin() as session:
        for values in rows:
            row = {
                header: _value(value)
                for header, value in zip(headers, values, strict=True)
            }
            if not any(value is not None for value in row.values()):
                continue
            absent = [name for name in REQUIRED if row.get(name) is None]
            if absent:
                raise SystemExit(
                    f"Row {ingested + 2} is missing required values: {', '.join(sorted(absent))}"
                )
            brand_name = str(row.get("brand") or "Maruti Suzuki")
            brand = session.scalar(
                select(Brand).where(func.lower(Brand.name) == brand_name.lower())
            )
            if not brand:
                brand = Brand(name=brand_name)
                session.add(brand)
                session.flush()
            model_name = str(row["model"])
            vehicle_model = session.scalar(
                select(VehicleModel).where(
                    VehicleModel.brand_id == brand.id,
                    func.lower(VehicleModel.name) == model_name.lower(),
                )
            )
            if not vehicle_model:
                vehicle_model = VehicleModel(brand_id=brand.id, name=model_name)
                session.add(vehicle_model)
                session.flush()
            variant_statement = insert(Variant).values(
                catalogue_id=str(row["catalogue_id"]),
                model_id=vehicle_model.id,
                trim=str(row["trim"]),
                variant_name=str(row["variant_name"]),
                model_year=int(row["model_year"]),
                is_active=bool(row.get("is_active", True)),
            )
            variant_id = session.scalar(
                variant_statement.on_conflict_do_update(
                    index_elements=[Variant.catalogue_id],
                    set_={
                        "model_id": vehicle_model.id,
                        "trim": str(row["trim"]),
                        "variant_name": str(row["variant_name"]),
                        "model_year": int(row["model_year"]),
                        "is_active": bool(row.get("is_active", True)),
                    },
                ).returning(Variant.id)
            )
            session.execute(
                insert(VariantPrice)
                .values(
                    variant_id=variant_id,
                    city=str(row["city"]),
                    ex_showroom_price=row["ex_showroom_price"],
                    on_road_price=row.get("on_road_price"),
                )
                .on_conflict_do_update(
                    index_elements=[VariantPrice.variant_id, VariantPrice.city],
                    set_={
                        "ex_showroom_price": row["ex_showroom_price"],
                        "on_road_price": row.get("on_road_price"),
                    },
                )
            )
            for spec_model in SPEC_MODELS:
                _upsert_spec(session, spec_model, int(variant_id), row)
            seen.add(str(row["catalogue_id"]))
            ingested += 1
        if args.deactivate_missing and seen:
            session.execute(
                Variant.__table__.update()
                .where(Variant.catalogue_id.not_in(seen))
                .values(is_active=False)
            )
    print(f"Ingested {ingested} catalogue rows from sheet {sheet.title!r}.")


if __name__ == "__main__":
    main()
